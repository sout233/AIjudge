import json
import os
import threading
from datetime import datetime
from typing import Any, Optional

from fastapi import HTTPException
from jinja2 import Environment, FileSystemLoader
from loguru import logger

from app.config.config import REPORT_TEMPLATE_DIR, RESULT_DIR

PDF_CACHE_DIR = os.path.join(RESULT_DIR, "pdf_cache")
os.makedirs(PDF_CACHE_DIR, exist_ok=True)
os.makedirs(REPORT_TEMPLATE_DIR, exist_ok=True)

_jinja_env: Optional[Environment] = None
PDF_LOCKS: dict[str, threading.Lock] = {}
_weasyprint_available = False


def _get_jinja_env() -> Environment:
    """Lazily initialize the Jinja environment."""
    global _jinja_env
    if _jinja_env is None:
        _jinja_env = Environment(loader=FileSystemLoader(searchpath=REPORT_TEMPLATE_DIR))
    return _jinja_env


def _check_weasyprint() -> bool:
    """Check whether WeasyPrint is available."""
    global _weasyprint_available
    if not _weasyprint_available:
        try:
            from weasyprint import HTML  # noqa: F401

            _weasyprint_available = True
        except ImportError:
            logger.error("weasyprint 未安装或系统依赖缺失")
    return _weasyprint_available


def _load_score_result(raw_result: Any) -> dict[str, Any]:
    """Parse Dify output and unwrap the outer `result` field when present."""
    if isinstance(raw_result, str):
        try:
            raw_result = json.loads(raw_result)
        except json.JSONDecodeError as exc:
            raise HTTPException(400, detail="评分结果不是合法 JSON") from exc

    if not isinstance(raw_result, dict):
        raise HTTPException(400, detail="评分结果格式不正确")

    normalized = raw_result.get("result", raw_result)
    if not isinstance(normalized, dict):
        raise HTTPException(400, detail="评分结果内容格式不正确")

    return normalized


def _build_report_context(score_result: dict[str, Any]) -> dict[str, Any]:
    """Prepare template context for single-judge and multi-judge reports."""
    evaluations = score_result.get("evaluations")
    is_multi_judge = isinstance(evaluations, list) and len(evaluations) > 0

    if is_multi_judge:
        valid_evaluations = [evaluation for evaluation in evaluations if isinstance(evaluation, dict)]
        scores = [int(evaluation.get("total_score", 0)) for evaluation in valid_evaluations]
        average_score = round(sum(scores) / len(scores)) if scores else 0
        final_review = score_result.get("final_review")
        final_review = final_review if isinstance(final_review, dict) else {}
        fallback_max_score = int(valid_evaluations[0].get("max_score", 100)) if valid_evaluations else 100

        return {
            "result": score_result,
            "is_multi_judge": True,
            "summary_score": int(final_review.get("final_total_score", average_score)),
            "summary_max_score": int(final_review.get("final_max_score", fallback_max_score)),
            "average_score": average_score,
            "final_comment": final_review.get("final_comment", ""),
            "score_reason": final_review.get("score_reason", ""),
            "highest_score": max(scores) if scores else 0,
            "lowest_score": min(scores) if scores else 0,
            "judge_count": len(valid_evaluations),
        }

    total_score = int(score_result.get("total_score", 0))
    max_score = int(score_result.get("max_score", 100))
    return {
        "result": score_result,
        "is_multi_judge": False,
        "summary_score": total_score,
        "summary_max_score": max_score,
        "average_score": total_score,
        "final_comment": "",
        "score_reason": "",
        "highest_score": total_score,
        "lowest_score": total_score,
        "judge_count": 1,
    }


def generate_pdf_sync(workflow_run_id: str) -> str:
    """Generate a PDF report for a workflow run."""
    pdf_file_path = os.path.join(PDF_CACHE_DIR, f"{workflow_run_id}.pdf")

    if os.path.exists(pdf_file_path):
        return pdf_file_path

    if not _check_weasyprint():
        raise HTTPException(500, detail="PDF 生成功能暂不可用，请联系管理员")

    lock = PDF_LOCKS.setdefault(workflow_run_id, threading.Lock())
    with lock:
        if os.path.exists(pdf_file_path):
            return pdf_file_path

        try:
            result_path = os.path.join(RESULT_DIR, f"{workflow_run_id}.json")
            if not os.path.exists(result_path):
                raise HTTPException(404, detail="报告文件不存在")

            with open(result_path, "r", encoding="utf-8") as file:
                data = json.load(file)

            outputs = data.get("workflow_data", {}).get("data", {}).get("outputs", {})
            raw_score_result = outputs.get("result")
            if raw_score_result is None:
                raw_score_result = outputs.get("text")
            if not raw_score_result:
                raise HTTPException(400, detail="评分结果尚未生成")

            score_result = _load_score_result(raw_score_result)
            report_context = _build_report_context(score_result)

            workflow_data = data.get("workflow_data", {})
            report_id = workflow_data.get("workflow_run_id", workflow_run_id)
            finished_at_timestamp = workflow_data.get("data", {}).get("finished_at")
            generate_time = (
                datetime.fromtimestamp(finished_at_timestamp).strftime("%Y-%m-%d %H:%M:%S")
                if finished_at_timestamp
                else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )

            template_file = os.path.join(REPORT_TEMPLATE_DIR, "report_template.html")
            if not os.path.exists(template_file):
                raise HTTPException(500, detail="PDF 模板文件缺失")

            jinja_env = _get_jinja_env()
            template = jinja_env.get_template("report_template.html")
            html_content = template.render(
                **report_context,
                report_id=report_id,
                generate_time=generate_time,
            )

            from weasyprint import HTML

            HTML(string=html_content).write_pdf(pdf_file_path)
        except HTTPException:
            if os.path.exists(pdf_file_path):
                os.remove(pdf_file_path)
            raise
        except Exception as exc:
            if os.path.exists(pdf_file_path):
                os.remove(pdf_file_path)
            logger.exception(f"PDF 生成失败 [{workflow_run_id}]: {exc}")
            raise HTTPException(500, detail=f"PDF 生成失败: {type(exc).__name__}: {exc}") from exc

    return pdf_file_path


async def export_zip_batch_results(manifest_id: str):
    """导出 ZIP 批量任务结果为 ZIP 文件（包含所有 PDF 和 Excel 汇总表）"""
    import io
    import zipfile
    import tempfile
    import shutil
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.config.config import RESULT_DIR
    
    manifest_path = os.path.join(RESULT_DIR, f"zip_manifest_{manifest_id}.json")
    if not os.path.exists(manifest_path):
        raise HTTPException(404, "任务清单不存在")
    
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = json.load(f)
    
    tasks = manifest.get("tasks", [])
    logger.info(f"开始导出ZIP，共 {len(tasks)} 个任务")
    
    # 使用临时文件夹
    with tempfile.TemporaryDirectory() as temp_dir:
        # 创建 reports 目录
        reports_dir = os.path.join(temp_dir, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        logger.info(f"创建临时目录: {reports_dir}")
        
        # 1. 批量生成PDF并复制到临时目录（使用多线程并发，最多4个并发）
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        
        pdf_copied_count = 0
        pdf_failed_count = 0
        counter_lock = threading.Lock()
        
        def process_single_task(task: dict) -> bool:
            """处理单个任务，返回是否成功"""
            nonlocal pdf_copied_count, pdf_failed_count
            
            workflow_run_id = task.get("workflow_run_id")
            original_name = task.get("original_name", task.get("filename", "未知文件"))
            
            if not workflow_run_id:
                logger.warning(f"任务缺少 workflow_run_id: {task}")
                return False
            
            # 构建结果文件路径（使用 RESULT_DIR）
            result_path = os.path.join(RESULT_DIR, f"{workflow_run_id}.json")
            
            if not os.path.exists(result_path):
                logger.warning(f"结果文件不存在: {result_path}")
                return False
            
            try:
                with open(result_path, "r", encoding="utf-8") as rf:
                    result_data = json.load(rf)
                
                status = result_data.get("status", "")
                
                # 只有已完成的任务才生成PDF
                if status in ["success", "succeeded", "error", "failed"]:
                    try:
                        # 生成PDF
                        pdf_path = generate_pdf_sync(workflow_run_id)
                        logger.info(f"PDF已生成: {pdf_path}")
                        
                        # 复制到临时 reports 目录
                        safe_name = "".join(c for c in original_name if c.isalnum() or c in '._- ').strip()
                        if not safe_name:
                            safe_name = f"report_{workflow_run_id[:8]}"
                        
                        dest_path = os.path.join(reports_dir, f"{safe_name}_{workflow_run_id[:8]}.pdf")
                        shutil.copy2(pdf_path, dest_path)
                        logger.info(f"PDF已复制到: {dest_path}")
                        
                        with counter_lock:
                            nonlocal pdf_copied_count
                            pdf_copied_count += 1
                        return True
                    except Exception as e:
                        logger.warning(f"生成/复制PDF失败 {workflow_run_id}: {e}")
                        with counter_lock:
                            nonlocal pdf_failed_count
                            pdf_failed_count += 1
                        return False
                else:
                    logger.info(f"任务状态不适合生成PDF: {workflow_run_id}, 状态: {status}")
                    return False
            except Exception as e:
                logger.warning(f"处理任务失败 {workflow_run_id}: {e}")
                return False
        
        # 收集需要处理的任务
        tasks_to_process = [task for task in tasks if task.get("workflow_run_id")]
        total_tasks = len(tasks_to_process)
        
        if total_tasks > 0:
            logger.info(f"开始并发生成PDF，共 {total_tasks} 个任务，最大并发数: 4")
            
            # 使用线程池并发处理（最多4个并发）
            with ThreadPoolExecutor(max_workers=4) as executor:
                # 提交所有任务
                future_to_task = {
                    executor.submit(process_single_task, task): task 
                    for task in tasks_to_process
                }
                
                # 等待所有任务完成
                for future in as_completed(future_to_task):
                    task = future_to_task[future]
                    workflow_run_id = task.get("workflow_run_id", "unknown")
                    try:
                        future.result()
                    except Exception as e:
                        logger.error(f"任务执行异常 {workflow_run_id}: {e}")
        
        logger.info(f"PDF处理完成: 成功复制 {pdf_copied_count}, 失败 {pdf_failed_count}")
        
        # 2. 生成 Excel 汇总表
        wb = Workbook()
        ws = wb.active
        ws.title = "评审结果汇总"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 45  # 项目名称
        ws.column_dimensions['B'].width = 12  # 状态
        ws.column_dimensions['C'].width = 12  # 总分
        
        # 标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 写入表头（三列）
        headers = ["项目名称", "状态", "总分"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        
        # 准备收集数据
        completed_count = 0
        failed_count = 0
        total_score_sum = 0
        scored_count = 0
        
        # 写入数据行
        for task in tasks:
            original_name = task.get("original_name", task.get("filename", "未知文件"))
            workflow_run_id = task.get("workflow_run_id")
            
            # 使用 workflow_run_id 构建结果文件路径
            result_path = os.path.join(RESULT_DIR, f"{workflow_run_id}.json") if workflow_run_id else ""
            
            # 初始化行数据
            # 状态只有"成功"或"失败"两种，其他状态留空
            status_text = ""  
            score_value = None  # 总分数字，未完成或失败为空
            
            if result_path and os.path.exists(result_path):
                try:
                    with open(result_path, "r", encoding="utf-8") as rf:
                        result_data = json.load(rf)
                    
                    status = result_data.get("status", "unknown")
                    
                    if status in ["success", "succeeded"]:
                        status_text = "成功"
                        completed_count += 1
                        
                        # 提取分数
                        score, max_score = _extract_score_from_result_data(result_data)
                        if score is not None:
                            score_value = score
                            total_score_sum += score
                            scored_count += 1
                    elif status in ["error", "failed"]:
                        status_text = "失败"
                        failed_count += 1
                    # 其他状态（进行中、等待中等）保持空字符串
                        
                except Exception:
                    status_text = "失败"  # 读取异常也算失败
            
            # 写入行数据（三列）
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=original_name)
            ws.cell(row=row, column=2, value=status_text).alignment = Alignment(horizontal="center")
            if score_value is not None:
                ws.cell(row=row, column=3, value=score_value).alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=row, column=3, value="").alignment = Alignment(horizontal="center")
            
            # 根据状态设置背景色（三列）
            if status_text == "成功":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status_text == "失败":
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                fill = None
            
            if fill:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = fill
        
        # 添加统计信息行（放在第1列下方）
        stats_row = ws.max_row + 2
        ws.cell(row=stats_row, column=1, value="统计信息").font = Font(bold=True)
        ws.cell(row=stats_row + 1, column=1, value=f"完成任务数: {completed_count}")
        ws.cell(row=stats_row + 2, column=1, value=f"失败任务数: {failed_count}")
        ws.cell(row=stats_row + 3, column=1, value=f"总任务数: {len(tasks)}")
        if scored_count > 0:
            avg_score = round(total_score_sum / scored_count, 2)
            ws.cell(row=stats_row + 4, column=1, value=f"平均分数: {avg_score}")
        
        # 保存 Excel 到临时目录
        excel_path = os.path.join(temp_dir, "评审结果汇总.xlsx")
        wb.save(excel_path)
        logger.info(f"Excel已保存到: {excel_path}")
        
        # 添加说明文件
        readme_content = f"""ZIP 批量评审结果导出
{'=' * 50}

导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
任务清单 ID: {manifest_id}

文件说明:
- 评审结果汇总.xlsx: 包含所有文件的评审结果汇总表
- reports/: 包含所有 PDF 格式的评审报告

统计:
- 总任务数: {len(tasks)}
- 已完成: {completed_count}
- 失败: {failed_count}
- PDF 报告数: {pdf_copied_count}
"""
        readme_path = os.path.join(temp_dir, "README.txt")
        with open(readme_path, "w", encoding="utf-8") as f:
            f.write(readme_content)
        
        # 打包整个临时目录到ZIP
        logger.info(f"开始打包ZIP，临时目录: {temp_dir}")
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, arcname)
                    logger.info(f"添加到ZIP: {arcname}")
        
        logger.info("ZIP打包完成")
    
    zip_buffer.seek(0)
    return zip_buffer


def _extract_score_from_result_data(result_data: dict) -> tuple[float | None, float | None]:
    """从结果数据中提取分数（供 export_zip_batch_results 使用）"""
    try:
        # 注意：这里的 workflow_data 是 Dify API 返回的 workflow_data，不是外层包裹
        # 路径: result_data -> workflow_data -> data -> outputs -> result
        workflow_data = result_data.get("workflow_data", {})
        data = workflow_data.get("data", {})
        outputs = data.get("outputs", {})
        
        # 尝试获取 result 或 text 字段
        raw_result = outputs.get("result") or outputs.get("text")
        if not raw_result:
            return None, None
        
        # 如果是字符串，尝试解析 JSON
        if isinstance(raw_result, str):
            try:
                parsed = json.loads(raw_result)
            except json.JSONDecodeError:
                return None, None
        else:
            parsed = raw_result
        
        # 处理 WrappedJudgeResult (包装格式)
        if isinstance(parsed, dict) and "result" in parsed:
            parsed = parsed["result"]
        
        # 多评审格式 (MultiJudgeResult)
        if isinstance(parsed, dict) and "evaluations" in parsed and isinstance(parsed["evaluations"], list):
            evaluations = parsed["evaluations"]
            if not evaluations:
                return None, None
            
            # 如果有最终评审结果，使用最终评审
            final_review = parsed.get("final_review")
            if final_review and isinstance(final_review, dict):
                return (
                    float(final_review.get("final_total_score", 0)),
                    float(final_review.get("final_max_score", 100))
                )
            
            # 否则计算平均分
            total_score = sum(e.get("total_score", 0) for e in evaluations)
            avg_score = round(total_score / len(evaluations))
            max_score = evaluations[0].get("max_score", 100) if evaluations else 100
            return float(avg_score), float(max_score)
        
        # 单评审格式 (JudgeResult)
        if isinstance(parsed, dict) and "total_score" in parsed:
            return (
                float(parsed.get("total_score", 0)),
                float(parsed.get("max_score", 100))
            )
        
        return None, None
    except Exception:
        return None, None

